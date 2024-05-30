import csv
import shutil
import openpyxl
import os
import sys
from openpyxl.utils import get_column_letter
from datetime import datetime


import random

def generate_random_numbers(n, m):
    # Generate n random numbers
    numbers = [random.uniform(0, 1) for _ in range(n)]
    
    # Normalize the numbers so their sum is m
    s = sum(numbers)
    numbers = [i/s * m for i in numbers]
    
    return numbers

def csv_to_json(csv_file):
    """
    Convert a CSV file to a list of dictionaries.
    Each dictionary represents a row in the CSV, with keys as column names and values as cell values.
    """
    with open(csv_file, 'r', encoding='utf-8-sig') as file:
        return list(csv.DictReader(file))
    
def filter_orders_by_customer(orders, customer_name):
    """
    Filter a list of orders to include only those for a specific customer.
    """
    return [order for order in orders if order['customerJoin_entityId0_searchValue'] == customer_name]

sales_orders = csv_to_json('daily-data/open_sales_orders.csv')

# Filter sales orders for a specific customer
sales_orders = filter_orders_by_customer(sales_orders, 'Target')

shipments_results = csv_to_json('daily-data/target/target_shipments_results.csv')

item_list = csv_to_json('daily-data/item_list.csv')

# Now we group the sales_orders by the basic_otherRefNum0_searchValue's last 4 digits
grouped_sales_orders = {}
for order in sales_orders:
    store_num = order['basic_otherRefNum0_searchValue'][-4:]
    if store_num not in grouped_sales_orders:
        grouped_sales_orders[store_num] = []
    grouped_sales_orders[store_num].append(order)

# now we add another dictionary to the grouped_sales_orders where the key is the "shipments_results" and the value is the shipment_result list, where shipment_result's Purchase Order Number matches basic_otherRefNum0_searchValue
grouped_shipments_results = {}
for shipment in shipments_results:
    store_num = shipment['Purchase Order Number'][-4:]
    if store_num not in grouped_shipments_results:
        grouped_shipments_results[store_num] = []
    grouped_shipments_results[store_num].append(shipment)

# print(grouped_shipments_results['0590'])

grouped_item_list = {}
for key in grouped_sales_orders.keys():
    for order in grouped_sales_orders[key]:
        for item in item_list:
            if item['Name'] == order['itemJoin_itemId0_searchValue']:
                if key not in grouped_item_list:
                    grouped_item_list[key] = []
                grouped_item_list[key].append(item)

# now we merge the grouped_sales_orders and grouped_shipments_results into a new dictionary called grouped_sales_orders_with_shipments_results where the key is the store number and the value is a dictionary with the key 'sales_orders' and 'shipments_results'
grouped_sales_orders_with_shipments_results = {}
for key in grouped_sales_orders.keys():
    grouped_sales_orders_with_shipments_results[key] = {
        'sales_orders': grouped_sales_orders[key], 
        'shipments_results': grouped_shipments_results[key], 
        'item_list': grouped_item_list[key]
    }
# print(grouped_sales_orders_with_shipments_results['0590'])
# Now we can write the grouped_routing_status_by_dest to the excel file
# we copy the template.xlsx to the new file with the name 'BOL-{PO Dest}.xlsx'

# python code to create a folder with today's date with format `yyyy-mm-dd`
today = datetime.now().strftime('%Y-%m-%d')
folder_name = "target-{}".format(today)
if not os.path.exists(folder_name):
    os.mkdir(folder_name)

print("There are {} store numbers so should have the same amount of BOL files".format(len(grouped_sales_orders_with_shipments_results.keys())))
for key in grouped_sales_orders_with_shipments_results.keys():
    single_sales_orders_with_shipments_results_dict = grouped_sales_orders_with_shipments_results.get(key, {})
    # print(single_sales_orders_with_shipments_results_dict)
    file_name = 'BOL-{}-target.xlsx'.format(key)
    file_path = "{}/{}".format(folder_name, file_name)
    shutil.copy('template/template-target.xlsx', file_path)
    wb = openpyxl.load_workbook(file_path)
    ws = wb['Sheet1']
    # Insert spreadsheet's sheet1's B1 cell with the value today's date with format `mm/dd/yyyy`
    ws['B1'] = datetime.now().strftime('%m/%d/%Y')
    # Insert spreadsheet's sheet1's J2 cell with the routing_status's Load ID value
    
    shipments_results_list = single_sales_orders_with_shipments_results_dict.get('shipments_results', [])
    
    if len(shipments_results_list) > 0:
        ws['J2'] = shipments_results_list[0].get('Bill of Lading', '')
        # Insert spreadsheet's sheet1's J3 cell with the routing_status's Load ID value
        ws['J3'] = shipments_results_list[0].get('Target Dispatch', '')
        ws['J4'] = shipments_results_list[0].get('SECO Routing', '')
        

        # Insert spreadsheet's sheet1's H8 cell with the routing_status's Carrier Name value
        ws['H8'] = "CARRIER NAME: {}".format(shipments_results_list[0].get('Carrier', ''))
        # insert spreadsheet's sheet1's H11 cell with the routing_status's Industry SCAC value
        ws['H11'] = "SCAC: {}".format(shipments_results_list[0].get('SCAC', ''))
        
        # Insert spreadsheet's sheet1's A9 cell with the sales'order's shippingAddressJoin_addressee0_searchValue value
        sales_orders_list = single_sales_orders_with_shipments_results_dict.get('sales_orders', [])
        if len(sales_orders_list) > 0:
            ws['A9'] = sales_orders_list[0].get('shippingAddressJoin_addressee0_searchValue', '')
            # Insert spreadsheet's sheet1's A10 cell with the sales'order's shippingAddressJoin_address1_searchValue value
            ws['A10'] = sales_orders_list[0].get('shippingAddressJoin_address10_searchValue', '')
            # Insert spreadsheet's sheet1's A11 cell with the sales'order's shippingAddressJoin_city0_searchValue, shippingAddressJoin_city0_searchValue shippingAddressJoin_zip0_searchValue  combined value
            ws['A11'] = '{}, {} {}'.format(sales_orders_list[0].get('shippingAddressJoin_city0_searchValue', ''), sales_orders_list[0].get('shippingAddressJoin_state0_searchValue', ''), sales_orders_list[0].get('shippingAddressJoin_zip0_searchValue', ''))
        
        
        row = 24
        for rs in shipments_results_list:
            ws['{}{}'.format('B', row)] = rs.get('Purchase Order Number', '')
            ws['{}{}'.format('F', row)] = int(rs.get('Packages', ''))
            ws['{}{}'.format('G', row)] = int(rs.get('Weight', ''))
            row += 1

        grouped_item_list = single_sales_orders_with_shipments_results_dict.get('item_list', [])
        row = 32

        random_weights = generate_random_numbers(len(sales_orders_list), int(shipments_results_list[0].get('Weight', '')))    
        
        for sales_order in sales_orders_list:
            handle_unit_quantity = int(sales_order['basic_quantity0_searchValue'])
            item_cases = 0
            display_name = sales_order['itemJoin_itemId0_searchValue']
            for item in grouped_item_list:
                if item.get('Name', '') == sales_order['itemJoin_itemId0_searchValue']:
                    display_name += " {}".format(item.get('Display Name', '')) 
                    item_cases = int(handle_unit_quantity / int(item['Package Quantity']))

            ws['{}{}'.format('A', row)] = handle_unit_quantity
            ws['{}{}'.format('B', row)] = 'Units'
            ws['{}{}'.format('C', row)] = item_cases
            ws['{}{}'.format('D', row)] = 'Carton'
            ws['{}{}'.format('E', row)] = random_weights[row - 32] 
            ws['{}{}'.format('G', row)] = display_name
            
            row += 1

        

    
    wb.save(file_path)
    
    

    