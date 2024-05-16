import csv
import shutil
import openpyxl
import os
import sys
from openpyxl.utils import get_column_letter
from datetime import datetime


def csv_to_json(csv_file):
    """
    Convert a CSV file to a list of dictionaries.
    Each dictionary represents a row in the CSV, with keys as column names and values as cell values.
    """
    with open(csv_file, 'r', encoding='utf-8-sig') as file:
        return list(csv.DictReader(file))
    
def filter_orders_by_customer(orders, customer_name):
    """
    Filter a list of orders to include only those for a specific customer.
    """
    return [order for order in orders if order['customerJoin_entityId0_searchValue'] == customer_name]

sales_orders = csv_to_json('daily-data/open_sales_orders.csv')

# Filter sales orders for a specific customer
sales_orders = filter_orders_by_customer(sales_orders, 'Wal-Mart')

routing_status = csv_to_json('daily-data/routing_status.csv')

item_list = csv_to_json('daily-data/item_list.csv')


# Now we group the routing_status by the routing status's PO Dest
grouped_routing_status_by_dest = {}
for status in routing_status:
    po_dest = status['PO Dest']  # Assuming 'PO Dest' is the column name in your CSV
    if po_dest not in grouped_routing_status_by_dest:
        grouped_routing_status_by_dest[po_dest] = []
    grouped_routing_status_by_dest[po_dest].append(status)

# Now we turn the grouped_routing_status_by_dest into a dictionary where the key is the PO Dest and the value is dictionary with the key 'routing_status' and the value is the routing_status list
grouped_routing_status_by_dest = {key: {'routing_status': value} for key, value in grouped_routing_status_by_dest.items()}
# print(grouped_routing_status_by_dest)

empty_order_po_num_for_dest = []
# Append the sales orders to the grouped_routing_status_by_dest where the routing_status's PO Number matches basic_otherRefNum0_searchValue
for status in routing_status:
    single_grouped_routing_status_dict = grouped_routing_status_by_dest.get(status['PO Dest'], {})
    single_grouped_routing_status = single_grouped_routing_status_dict.get('routing_status', [])
    grouped_sales_order = []
    grouped_item_list = []
    for order in sales_orders:
        for single_status in single_grouped_routing_status:
            if order['basic_otherRefNum0_searchValue'] == single_status['PO Number']:
                grouped_sales_order.append(order)

    if len(grouped_sales_order) == 0:
        empty_order_po_num_for_dest.append(status['PO Dest'])
        print("No PO Number {} found in Open Sales Orders. Grouped Sales Order is empty for PO Dest: {}. ".format(single_status['PO Number'], status['PO Dest']))
        # sys.exit()
    single_grouped_routing_status_dict['sales_orders'] = grouped_sales_order    
    
    for order in grouped_sales_order:    
        for item in item_list:
            if item['Name'] == order['itemJoin_itemId0_searchValue']:
                grouped_item_list.append(item)
    
    single_grouped_routing_status_dict['item_list'] = grouped_item_list
    
    grouped_routing_status_by_dest[status['PO Dest']] = single_grouped_routing_status_dict
            
# Now we can write the grouped_routing_status_by_dest to the excel file
# we copy the template.xlsx to the new file with the name 'BOL-{PO Dest}.xlsx'

# python code to create a folder with today's date with format `yyyy-mm-dd`
today = datetime.now().strftime('%Y-%m-%d')
folder_name = "walmart-{}".format(today)
if not os.path.exists(folder_name):
    os.mkdir(folder_name)

print("There are {} destinations so should have the same amount of BOL files".format(len(grouped_routing_status_by_dest.keys())))
for key in grouped_routing_status_by_dest.keys():
    # key = '6054'
    if key in empty_order_po_num_for_dest:
        continue
    single_grouped_routing_status_dict = grouped_routing_status_by_dest.get(key, {})
    file_name = 'BOL-{}.xlsx'.format(key)
    file_path = "{}/{}".format(folder_name, file_name)
    shutil.copy('template/template-walmart.xlsx', file_path)
    wb = openpyxl.load_workbook(file_path)
    ws = wb['Sheet1']
    # Insert spreadsheet's sheet1's B1 cell with the value today's date with format `mm/dd/yyyy`
    ws['B1'] = datetime.now().strftime('%m/%d/%Y')
    # Insert spreadsheet's sheet1's J2 cell with the routing_status's Load ID value
    
    routing_status_list = single_grouped_routing_status_dict.get('routing_status', [])
    
    if len(routing_status_list) > 0:
        ws['J2'] = routing_status_list[0].get('Load ID', '')
        # Insert spreadsheet's sheet1's J3 cell with the routing_status's Load ID value
        ws['J3'] = routing_status_list[0].get('Load ID', '')

        # Insert spreadsheet's sheet1's H8 cell with the routing_status's Carrier Name value
        ws['H8'] = "CARRIER NAME: {}".format(routing_status_list[0].get('Carrier Name', ''))
        # insert spreadsheet's sheet1's H11 cell with the routing_status's Industry SCAC value
        ws['H11'] = "SCAC: {}".format(routing_status_list[0].get('Industry SCAC', ''))
        
        # Insert spreadsheet's sheet1's A9 cell with the sales'order's shippingAddressJoin_addressee0_searchValue value
        sales_orders_list = single_grouped_routing_status_dict.get('sales_orders', [])
        if len(sales_orders_list) > 0:
            ws['A9'] = sales_orders_list[0].get('shippingAddressJoin_addressee0_searchValue', '')
            # Insert spreadsheet's sheet1's A10 cell with the sales'order's shippingAddressJoin_address1_searchValue value
            ws['A10'] = sales_orders_list[0].get('shippingAddressJoin_address10_searchValue', '')
            # Insert spreadsheet's sheet1's A11 cell with the sales'order's shippingAddressJoin_city0_searchValue, shippingAddressJoin_city0_searchValue shippingAddressJoin_zip0_searchValue  combined value
            ws['A11'] = '{}, {} {}'.format(sales_orders_list[0].get('shippingAddressJoin_city0_searchValue', ''), sales_orders_list[0].get('shippingAddressJoin_state0_searchValue', ''), sales_orders_list[0].get('shippingAddressJoin_zip0_searchValue', ''))
        
        
        row = 23
        for rs in routing_status_list:
            ws['{}{}'.format('A', row)] = rs.get('PO Number', '')
            total_weight_by_po_number = 0
            for sales_order in sales_orders_list:
                if rs.get('PO Number', '') == sales_order.get('basic_otherRefNum0_searchValue', ''):
                    for item in grouped_item_list:
                        if item.get('Name', '') == sales_order.get('itemJoin_itemId0_searchValue', ''):
                            total_weight_by_po_number += float(sales_order.get('itemJoin_weight0_searchValue', '')) *int(sales_order['basic_quantity0_searchValue'])
        
            ws['{}{}'.format('D', row)] = int(rs.get('Cases', ''))
            ws['{}{}'.format('E', row)] = total_weight_by_po_number #rs.get('Weight', '')
            ws['{}{}'.format('G', row)] = rs.get('MABD', '')
            ws['B1'] = rs.get('MABD', '')
            ws['{}{}'.format('H', row)] = rs.get('PO Dest', '')
            ws['{}{}'.format('I', row)] = rs.get('PO Type', '')
            ws['{}{}'.format('J', row)] = rs.get('Department', '')
            row += 1

        grouped_item_list = single_grouped_routing_status_dict.get('item_list', [])
        grouped_sales_order = single_grouped_routing_status_dict.get('sales_orders', [])
        row = 32

        for item in grouped_item_list:
            item_name = item.get('Name', '')
            # print(item_name)
            handle_unit_quantity = 0
            weight = 0
            # line_cube = 0
            item_cases = 0
            weight_by_po_number = {}
            for sales_order in grouped_sales_order:
                if item_name == sales_order['itemJoin_itemId0_searchValue']:
                    handle_unit_quantity = int(sales_order['basic_quantity0_searchValue']) #should be correct
                    item_cases = int(handle_unit_quantity / int(item['Package Quantity']))
                    weight =  float(sales_order['itemJoin_weight0_searchValue'])*int(sales_order['basic_quantity0_searchValue'])
                    # line_cube = sales_order['custitem_height']*sales_order['custitem_length']*sales_order['custitem_width']/1728*sales_order['basic_quantity0_searchValue']/sales_order['custitem_hj_tc_autopackquantity']
                    break
            ws['{}{}'.format('A', row)] = int(handle_unit_quantity)
            ws['{}{}'.format('B', row)] = 'Units'
            ws['{}{}'.format('C', row)] = int(item_cases)
            ws['{}{}'.format('D', row)] = 'Carton'
            ws['{}{}'.format('E', row)] =  weight 
            ws['{}{}'.format('G', row)] = "{} {}".format(item.get('Name', ''),item.get('Display Name', ''))
        
            row += 1

    
    wb.save(file_path)

    