import csv
import shutil
import openpyxl
import os
import sys
from openpyxl.utils import get_column_letter
from datetime import datetime


def csv_to_json(csv_file):
    """
    Convert a CSV file to a list of dictionaries.
    Each dictionary represents a row in the CSV, with keys as column names and values as cell values.
    """
    with open(csv_file, 'r', encoding='utf-8-sig') as file:
        return list(csv.DictReader(file))
    
def filter_orders_by_customer(orders, customer_name):
    """
    Filter a list of orders to include only those for a specific customer.
    """
    return [order for order in orders if order['customerJoin_entityId0_searchValue'] == customer_name]

# summarize the total quantity, total cases, total weight, and description of the item by its name
def group_all_items_property_by_item_name(grouped_routing_status_by_dest):
    item_property_by_name = {}
    for key in grouped_routing_status_by_dest.keys():
        single_grouped_routing_status_dict = grouped_routing_status_by_dest.get(key, {})
        grouped_item_list = single_grouped_routing_status_dict.get('item_list', [])
        grouped_sales_order = single_grouped_routing_status_dict.get('sales_orders', [])
        
        for sales_order in grouped_sales_order:
            handle_unit_quantity = 0
            weight = 0
            item_cases = 0
            item_id = sales_order['itemJoin_itemId0_searchValue']
            
            if item_id not in item_property_by_name:
                item_property_by_name[item_id] = {'total_quantity': 0, 'total_cases': 0, 'total_weight': 0, 'description': ''}
            else: 
                for item in grouped_item_list:
                    # if item['Name'] == 'FG0300-13':
                    #     print("Item:{}".format(item))
                    #     print("Sales Order:{}".format(sales_order))
                    if item_id == item['Name']:
                        handle_unit_quantity = int(sales_order['basic_quantity0_searchValue'])
                        item_cases = int(handle_unit_quantity / int(item['Package Quantity']))
                        weight =  float(sales_order['itemJoin_weight0_searchValue'])*int(sales_order['basic_quantity0_searchValue'])
                        item_property_by_name[item_id]['total_quantity'] += handle_unit_quantity
                        item_property_by_name[item_id]['total_cases'] += item_cases
                        item_property_by_name[item_id]['total_weight'] += weight
                        item_property_by_name[item_id]['description'] = "{} {}".format(item.get('Name', ''),item.get('Display Name', ''))
                        break
                
                # print("{}:{}".format(item_id, item_property_by_name[item_id]))
        
    return item_property_by_name

def generate_master_bol_first_page(folder_name, grouped_routing_status_by_dest):
    # set up for the target's master BOL file
    master_bol_file_path = "{}/master-target-bol.xlsx".format(folder_name)
    shutil.copy('template/template-target-master-bol.xlsx', master_bol_file_path)
    wb_master_bol = openpyxl.load_workbook(master_bol_file_path)
    ws_master_bol = wb_master_bol['Sheet1']

    load_id = ''
    for key in grouped_routing_status_by_dest.keys():
        # key = '6054'
        single_grouped_routing_status_dict = grouped_routing_status_by_dest.get(key, {})
        routing_status_list = single_grouped_routing_status_dict.get('routing_status', [])
        if len(routing_status_list) > 0:

            # Insert spreadsheet's sheet1's B1 cell with the value today's date with format `mm/dd/yyyy`
            ws_master_bol['A1'] = "DATE: {}".format(routing_status_list[0].get('Carrier PU Date', ''))
            load_id = routing_status_list[0].get('Load ID', '')
            # Insert spreadsheet's sheet1's J2 cell with the routing_status's Load ID value
            ws_master_bol['F3'] = "LOAD NUMBER: ".format(load_id)

            ws_master_bol['F5'] = "CARRIER NAME: {}".format(routing_status_list[0].get('Carrier Name', ''))
            # insert spreadsheet's sheet1's H11 cell with the routing_status's Industry SCAC value
            ws_master_bol['F7'] = "SCAC: {}".format(routing_status_list[0].get('Industry SCAC', ''))

            sales_orders_list = single_grouped_routing_status_dict.get('sales_orders', [])
            if len(sales_orders_list) > 0:
                ws_master_bol['A12'] = sales_orders_list[0].get('shippingAddressJoin_addressee0_searchValue', '')
                # Insert spreadsheet's sheet1's A10 cell with the sales'order's shippingAddressJoin_address1_searchValue value
                ws_master_bol['A13'] = sales_orders_list[0].get('shippingAddressJoin_address10_searchValue', '')
                # Insert spreadsheet's sheet1's A11 cell with the sales'order's shippingAddressJoin_city0_searchValue, shippingAddressJoin_city0_searchValue shippingAddressJoin_zip0_searchValue  combined value
                ws_master_bol['A14'] = '{}, {} {}'.format(sales_orders_list[0].get('shippingAddressJoin_city0_searchValue', ''), sales_orders_list[0].get('shippingAddressJoin_state0_searchValue', ''), sales_orders_list[0].get('shippingAddressJoin_zip0_searchValue', ''))
            
        break

    
    item_property_by_name = group_all_items_property_by_item_name(grouped_routing_status_by_dest)

    row = 25
    for key in item_property_by_name.keys():
        item = item_property_by_name.get(key, {})
        handle_unit_quantity = item.get('total_quantity', 0)
        weight = item.get('total_weight', 0)
        item_cases = item.get('total_cases', 0)
        ws_master_bol['{}{}'.format('A', row)] = int(handle_unit_quantity)
        ws_master_bol['{}{}'.format('B', row)] = 'Unit'
        ws_master_bol['{}{}'.format('C', row)] = int(item_cases)
        ws_master_bol['{}{}'.format('D', row)] = 'Carton'
        ws_master_bol['{}{}'.format('E', row)] =  int(weight)
        print(item.get('description', ''))
        ws_master_bol['{}{}'.format('G', row)] = item.get('description', '')
    
        row += 1

    # set up the row 55 to the 
    row = 56
    for key in grouped_routing_status_by_dest.keys():
        single_grouped_routing_status_dict = grouped_routing_status_by_dest.get(key, {})
        routing_status_list = single_grouped_routing_status_dict.get('routing_status', [])
        num_of_pkgs = 0
        total_weight = 0
        for rs in routing_status_list:
            num_of_pkgs += int(rs.get('Cases', ''))
            total_weight += float(rs.get('Weight', ''))
            
        ws_master_bol['{}{}'.format('A', row)] = int(routing_status_list[0].get('PO Number', ''))
        ws_master_bol['{}{}'.format('C', row)] = num_of_pkgs
        ws_master_bol['{}{}'.format('D', row)] = total_weight
        ws_master_bol['{}{}'.format('E', row)] = 'Y/N'
        ws_master_bol['{}{}'.format('F', row)] = routing_status_list[0].get('MABD', '')
        ws_master_bol['{}{}'.format('G', row)] = int(routing_status_list[0].get('PO Dest', ''))
        ws_master_bol['{}{}'.format('H', row)] = int(routing_status_list[0].get('PO Type', ''))
        ws_master_bol['{}{}'.format('I', row)] = int(routing_status_list[0].get('Department', ''))
        
        row += 1

    wb_master_bol.save(master_bol_file_path)

    # rename this file to BOL-master-walmart-2nd-page-{Load Id}.xlsx
    new_master_bol_file_path = "{}/BOL-master-target-bol-{}.xlsx".format(folder_name, load_id)
    os.rename(master_bol_file_path, new_master_bol_file_path)

    return len(grouped_routing_status_by_dest.keys()) > 12     


# set up the 2nd page of target's master BOL file
def generate_master_bol_second_page(folder_name, grouped_routing_status_by_dest):
    master_bol_second_page_file_path = "{}/BOL-master-target-2nd-page.xlsx".format(folder_name)
    shutil.copy('template/template-target-master-bol-2nd-page.xlsx', master_bol_second_page_file_path)
    wb_master_bol_second_page = openpyxl.load_workbook(master_bol_second_page_file_path)
    ws_master_bol_second_page = wb_master_bol_second_page['Sheet1']

    if (len(grouped_routing_status_by_dest.keys()) <= 45):
        ws_master_bol_second_page['J1'] = 'Page 2 of 2'
    else:
        ws_master_bol_second_page['J1'] = 'Page 2 of 3'
        ws_master_bol_second_page['J52'] = 'Page 3 of 3'

    row = 5
    load_id = ''
    carrier_pu_date = ''
    print("There are {} destinations so should have the same amount rows".format(len(grouped_routing_status_by_dest.keys())))
    for key in grouped_routing_status_by_dest.keys():
        single_grouped_routing_status_dict = grouped_routing_status_by_dest.get(key, {})
        routing_status_list = single_grouped_routing_status_dict.get('routing_status', [])
        num_of_pkgs = 0
        total_weight = 0
        for rs in routing_status_list:
            load_id = rs.get('Load ID', '')
            carrier_pu_date = rs.get('Carrier PU Date', '')
            num_of_pkgs += int(rs.get('Cases', ''))
            total_weight += float(rs.get('Weight', ''))
            
        ws_master_bol_second_page['{}{}'.format('A', row)] = int(routing_status_list[0].get('PO Number', ''))
        ws_master_bol_second_page['{}{}'.format('C', row)] = num_of_pkgs
        ws_master_bol_second_page['{}{}'.format('D', row)] = total_weight
        ws_master_bol_second_page['{}{}'.format('E', row)] = 'Y/N'
        ws_master_bol_second_page['{}{}'.format('F', row)] = routing_status_list[0].get('MABD', '')
        ws_master_bol_second_page['{}{}'.format('G', row)] = int(routing_status_list[0].get('PO Dest', ''))
        ws_master_bol_second_page['{}{}'.format('H', row)] = int(routing_status_list[0].get('PO Type', ''))
        ws_master_bol_second_page['{}{}'.format('I', row)] = int(routing_status_list[0].get('Department', ''))
        
        row += 1

        #if row == 49, we will skip a few rows and jump to row 57
        if row == 49:
            row == 57
    
    ws_master_bol_second_page['A1'] = 'Date: {}'.format(carrier_pu_date)
    wb_master_bol_second_page.save(master_bol_second_page_file_path)

    # rename this file to BOL-master-walmart-2nd-page-{Load Id}.xlsx
    new_master_bol_second_page_file_path = "{}/BOL-master-target-2nd-page-{}.xlsx".format(folder_name, load_id)
    os.rename(master_bol_second_page_file_path, new_master_bol_second_page_file_path)


def setup():

    sales_orders = csv_to_json('daily-data/open_sales_orders.csv')

    # Filter sales orders for a specific customer
    sales_orders = filter_orders_by_customer(sales_orders, 'target')

    routing_status = csv_to_json('daily-data/target/walmart_routing_status.csv')

    item_list = csv_to_json('daily-data/item_list.csv')


    # Now we group the routing_status by the routing status's PO Dest
    grouped_routing_status_by_dest = {}
    for status in routing_status:
        po_dest = status['PO Dest']  # Assuming 'PO Dest' is the column name in your CSV
        if po_dest not in grouped_routing_status_by_dest:
            grouped_routing_status_by_dest[po_dest] = []
        grouped_routing_status_by_dest[po_dest].append(status)

    # Now we turn the grouped_routing_status_by_dest into a dictionary where the key is the PO Dest and the value is dictionary with the key 'routing_status' and the value is the routing_status list
    grouped_routing_status_by_dest = {key: {'routing_status': value} for key, value in grouped_routing_status_by_dest.items()}
    # print(grouped_routing_status_by_dest)

    empty_order_po_num_for_dest = []
    # Append the sales orders to the grouped_routing_status_by_dest where the routing_status's PO Number matches basic_otherRefNum0_searchValue
    for status in routing_status:
        single_grouped_routing_status_dict = grouped_routing_status_by_dest.get(status['PO Dest'], {})
        single_grouped_routing_status = single_grouped_routing_status_dict.get('routing_status', [])
        grouped_sales_order = []
        grouped_item_list = []
        for order in sales_orders:
            for single_status in single_grouped_routing_status:
                if order['basic_otherRefNum0_searchValue'] == single_status['PO Number']:
                    grouped_sales_order.append(order)

        if len(grouped_sales_order) == 0:
            empty_order_po_num_for_dest.append(status['PO Dest'])
            print("No PO Number {} found in Open Sales Orders. Grouped Sales Order is empty for PO Dest: {}. ".format(single_status['PO Number'], status['PO Dest']))
            # sys.exit()
        single_grouped_routing_status_dict['sales_orders'] = grouped_sales_order    
        
        for order in grouped_sales_order:    
            for item in item_list:
                if item['Name'] == order['itemJoin_itemId0_searchValue']:
                    grouped_item_list.append(item)
        
        single_grouped_routing_status_dict['item_list'] = grouped_item_list
        
        grouped_routing_status_by_dest[status['PO Dest']] = single_grouped_routing_status_dict
            
    return grouped_routing_status_by_dest


if __name__ == '__main__':
    grouped_routing_status_by_dest = setup()

    # result = group_all_items_property_by_item_name(grouped_routing_status_by_dest)
    # print(result)
    
    today = datetime.now().strftime('%Y-%m-%d')
    folder_name = "walmart-{}".format(today)
    if not os.path.exists(folder_name):
        os.mkdir(folder_name)
    
    if (generate_master_bol_first_page(folder_name, grouped_routing_status_by_dest)):
        generate_master_bol_second_page(folder_name, grouped_routing_status_by_dest)
    
    print("Done")
    

    